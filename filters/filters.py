import openpyxl
import statistics
import datetime

def read_data_from_excel(file, sheet_name):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]

    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        game = {
            'name': row[0],
            'image': row[1],
            'developer': row[2],
            'publisher': row[3],
            'genre': row[4],
            'style': row[5],
            'platform': row[6],
            'rating': row[7],
            'downloads': row[8],
            'release': row[9]
        }
        data.append(game)

    return data


def sort_data(data, time_period='year'):
    ratings = [game['rating'] for game in data]
    mean_rating = statistics.mean(ratings)
    
    sorted_data = [game for game in data if game['rating'] >= mean_rating]
    sorted_data = sorted(sorted_data, key=lambda x: x['rating'], reverse=True)
    sorted_data = sorted(sorted_data, key=lambda x: x['downloads'], reverse=True)

    current_date = datetime.date.today()
    sorted_data_final = []

    while len(sorted_data) > 0:
        rating_group = [game for game in sorted_data if game['rating'] == sorted_data[0]['rating']]
        sorted_data = [game for game in sorted_data if game['rating'] != sorted_data[0]['rating']]

        if len(rating_group) > 1:
            game_average_tuples = []

            for game in rating_group:
                release_date = datetime.datetime.strptime(game['release'], '%b %d, %Y').date()
                if time_period == 'year':
                    time_delta = current_date.year - release_date.year
                elif time_period == 'month':
                    time_delta = (current_date.year - release_date.year) * 12 + (current_date.month - release_date.month)
                else:
                    raise ValueError("Invalid time_period specified. Supported values: 'year', 'month'")

                adjusted_time_delta = time_delta + (2 * min(1, time_delta))

                average_downloads = float(game['downloads'].replace(',', '.')) / (adjusted_time_delta + 1)
                game_average_tuples.append((game, average_downloads))

            sorted_tuples = sorted(game_average_tuples, key=lambda x: x[1], reverse=True)
            sorted_games = [game for game, _ in sorted_tuples]
            sorted_data_final.extend(sorted_games)
        else:
            sorted_data_final.extend(rating_group)

    sorted_data_final.extend([game for game in sorted_data])

    return sorted_data_final


def save_sorted_data_to_excel(data, file):
    sorted_workbook = openpyxl.Workbook()
    sorted_sheet = sorted_workbook.active
    sorted_sheet.title = 'Sorted Games'

    header_row = ["name", "image", "developer", "publisher", "genre", "style",
                  "platform", "rating", "downloads", "release"]
    sorted_sheet.append(header_row)

    for game in data:
        row_data = [
            game['name'],
            game['image'],
            game['developer'],
            game['publisher'],
            game['genre'],
            game['style'],
            game['platform'],
            game['rating'],
            game['downloads'],
            game['release']
        ]
        sorted_sheet.append(row_data)
        
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
        sorted_sheet.column_dimensions[col].width = 50

    sorted_workbook.save(file)

if __name__ == '__main__':

    input_file = '../games.xlsx'
    output_file = '../sorted_games.xlsx'
    input_sheet_name = 'Games'

    data = read_data_from_excel(input_file, input_sheet_name)

    sorted_data = sort_data(data)

    save_sorted_data_to_excel(sorted_data, output_file)