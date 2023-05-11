import os
import openpyxl

game_genres = {
    "Board Games": "Realism",
    "Educational": "Stylization",
    "Card": "Stylization",
    "Family": "Stylization",
    "Fighting": "Stylization",
    "Sports": "Realism",
    "Racing": "Realism",
    "Massively Multiplayer": "Realism",
    "Platformer": "Stylization",
    "Arcade": "Stylization",
    "Puzzle": "Stylization",
    "Simulation": "Realism",
    "Casual": "Stylization",
    "Shooter": "Realism",
    "Strategy": "Realism",
    "RPG": "Realism",
    "Adventure": "Stylization",
    "Indie": "Stylization",
    "Action": "Realism"
}

platforms = {
    'PC': 'PC',
    'PlayStation 5': 'Console',
    'PlayStation 4': 'Console',
    'Xbox One': 'Console',
    'Xbox Series S/X': 'Console',
    'Nintendo Switch': 'Console',
    'iOS': 'Mobile',
    'Android': 'Mobile',
    'Nintendo 3DS': 'Console',
    'Nintendo DS': 'Console',
    'Nintendo DSi': 'Console',
    'macOS': 'PC',
    'Linux': 'PC',
    'Xbox 360': 'Console',
    'Xbox': 'Console',
    'PlayStation 3': 'Console',
    'PlayStation 2': 'Console',
    'PlayStation': 'Console',
    'PS Vita': 'Console',
    'PSP': 'Console',
    'Wii U': 'Console',
    'Wii': 'Console',
    'GameCube': 'Console',
    'Nintendo 64': 'Console',
    'Game Boy Advance': 'Console',
    'Game Boy Color': 'Console',
    'Game Boy': 'Console',
    'SNES': 'Console',
    'NES': 'Console',
    'Classic Macintosh': 'PC',
    'Apple II': 'PC',
    'Commodore / Amiga': 'PC',
    'Atari 7800': 'Console',
    'Atari 5200': 'Console',
    'Atari 2600': 'Console',
    'Atari Flashback': 'Console',
    'Atari 8-bit': 'Console',
    'Atari ST': 'Console',
    'Atari Lynx': 'Console',
    'Atari XEGS': 'Console',
    'Genesis': 'Console',
    'SEGA Saturn': 'Console',
    'SEGA CD': 'Console',
    'SEGA 32X': 'Console',
    'SEGA Master System': 'Console',
    'Dreamcast': 'Console',
    '3DO': 'Console',
    'Jaguar': 'Console',
    'Game Gear': 'Console',
    'Neo Geo': 'Console',
    'Web': 'PC'
}

class Game:

    def __ini__(self):
        self.name = None
        self.image = None #
        self.developer = None
        self.publisher = None
        self.genre = None
        self.style = None
        self.platform = None
        self.rating = None
        self.downloads = None
        self.release = None

    def save(self):
        file = 'games.xlsx'
        sheet_name = 'Games'

        if not os.path.isfile(file):
            workbook = openpyxl.Workbook()
            workbook.save(file)

        workbook = openpyxl.load_workbook(file)

        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            row_index = sheet.max_row + 1
        else:
            sheet = workbook.create_sheet(sheet_name)
            header_row = ["name", "image", "developer", "publisher", "genre", "style",
                        "platform", "rating", "downloads", "release"]
            sheet.append(header_row)
            row_index = 2

        new_row = [
            getattr(self, 'name', ''),
            getattr(self, 'image', ''),
            getattr(self, 'developer', ''),
            getattr(self, 'publisher', ''),
            getattr(self, 'genre', ''),
            getattr(self, 'style', ''),
            getattr(self, 'platform', ''),
            getattr(self, 'rating', 0),
            getattr(self, 'downloads', 0),
            getattr(self, 'release', '')
        ]

        sheet = workbook[sheet_name]
        
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
            sheet.column_dimensions[col].width = 50

        sheet.append(new_row)
        workbook.save(file)



